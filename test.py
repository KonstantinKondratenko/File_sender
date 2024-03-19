from openpyxl import load_workbook

wb1 = load_workbook('общая.xlsx')
wb2 = load_workbook('фкит.xlsx')

surname_idx1 = 1
surname_idx2 = 3

if __name__ == '__main__':
	sheet1 = wb1.active
	sheet2 = wb2.active

	for row in sheet1.iter_rows():
		for cell in row:
			if cell.value == 'фкит':
				row_data = [cell.value for cell in row]
				break

	for row in sheet2.iter_rows():
		if row[surname_idx1].value == row_data[surname_idx2]:
			break  # опять же валидация по фамилии - нашли фамилию - скипаем строку
		else:
			sheet2.append(row_data)

	wb2.save('table2.xlsx')
